from flask import flash, render_template, jsonify, request, send_file
from flask_login import login_required, current_user
import pytz
from app.reports import bp
from app.models import Sale, Product, SaleItem, db
from datetime import datetime, timedelta
import io
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from sqlalchemy.orm import selectinload
from sqlalchemy import func

from app.utils.timezone import convert_utc_to_user_timezone

def convert_local_to_utc(local_dt):
    """Convert local datetime to UTC datetime"""
    try:
        # Jika timezone sudah tersedia di app, gunakan timezone user
        from app.utils.timezone import get_user_timezone
        user_tz = get_user_timezone()
        
        # Localize the datetime to user's timezone then convert to UTC
        localized_dt = user_tz.localize(local_dt)
        utc_dt = localized_dt.astimezone(pytz.UTC)
        return utc_dt
    except:
        # Fallback: assume local time is server time
        # Calculate the UTC offset
        local_now = datetime.now()
        utc_now = datetime.utcnow()
        offset = local_now - utc_now
        return local_dt - offset

def convert_utc_to_local(utc_dt):
    """Convert UTC datetime to local datetime"""
    try:
        from app.utils.timezone import get_user_timezone, convert_utc_to_user_timezone
        return convert_utc_to_user_timezone(utc_dt)
    except:
        # Fallback simple conversion
        local_now = datetime.now()
        utc_now = datetime.utcnow()
        offset = local_now - utc_now
        return utc_dt + offset

@bp.route('/')
@login_required
def index():
    return render_template('reports/index.html')

@bp.route('/sales-report')
@login_required
def sales_report():
    """Sales report dengan filter dan data chart - FIXED FILTER"""
    # Filter parameters
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    # Build query
    query = Sale.query.filter_by(tenant_id=current_user.tenant_id)
    
    # Debug filter parameters
    print(f"🔍 FILTER DEBUG: start_date={start_date_str}, end_date={end_date_str}")
    
    if start_date_str:
        try:
            # Parse tanggal dan konversi ke UTC untuk filter
            start_date_local = datetime.strptime(start_date_str, '%Y-%m-%d')
            # Konversi ke UTC awal hari
            start_date_utc = convert_local_to_utc(start_date_local.replace(hour=0, minute=0, second=0, microsecond=0))
            query = query.filter(Sale.created_at >= start_date_utc)
            print(f"📅 Filter start_date: {start_date_local} -> {start_date_utc}")
        except ValueError as e:
            flash('Format tanggal mulai tidak valid', 'error')
            print(f"❌ Error parsing start_date: {e}")
    
    if end_date_str:
        try:
            # Parse tanggal dan konversi ke UTC untuk filter
            end_date_local = datetime.strptime(end_date_str, '%Y-%m-%d')
            # Konversi ke UTC akhir hari
            end_date_utc = convert_local_to_utc(end_date_local.replace(hour=23, minute=59, second=59, microsecond=999999))
            query = query.filter(Sale.created_at <= end_date_utc)
            print(f"📅 Filter end_date: {end_date_local} -> {end_date_utc}")
        except ValueError as e:
            flash('Format tanggal akhir tidak valid', 'error')
            print(f"❌ Error parsing end_date: {e}")
    
    # Gunakan selectinload untuk loading relationships
    sales = query.options(
        selectinload(Sale.customer),
        selectinload(Sale.user)
    ).order_by(Sale.created_at.desc()).all()
    
    # Convert semua timestamp ke local timezone untuk display
    for sale in sales:
        sale.local_created_at = convert_utc_to_user_timezone(sale.created_at)
    
    # Pre-calculate items count dengan query terpisah
    sale_ids = [sale.id for sale in sales]
    
    if sale_ids:
        # Query untuk mendapatkan jumlah items per sale
        items_count_query = db.session.query(
            SaleItem.sale_id,
            func.count(SaleItem.id).label('items_count')
        ).filter(SaleItem.sale_id.in_(sale_ids))\
         .group_by(SaleItem.sale_id).all()
        
        # Buat dictionary untuk mapping sale_id -> items_count
        items_count_map = {sale_id: count for sale_id, count in items_count_query}
        
        # Assign items_count ke setiap sale
        for sale in sales:
            sale.items_count = items_count_map.get(sale.id, 0)
    else:
        for sale in sales:
            sale.items_count = 0
    
    # Summary statistics
    total_sales = len(sales)
    total_revenue = sum(sale.total_amount for sale in sales)
    avg_sale = total_revenue / total_sales if total_sales else 0
    
    # Prepare chart data untuk template
    payment_data = {}
    hourly_data = [0] * 24
    
    for sale in sales:
        # Payment method data
        method = sale.payment_method
        payment_data[method] = payment_data.get(method, 0) + 1
        
        # Hourly data - gunakan local time untuk chart
        try:
            # Gunakan local time hour untuk chart
            hour = sale.local_created_at.hour
            hourly_data[hour] += 1
        except:
            continue
    
    # Debug hasil filter
    print(f"📊 FILTER RESULT: {total_sales} sales found")
    
    return render_template('reports/sales_report.html', 
                         sales=sales,
                         total_sales=total_sales,
                         total_revenue=total_revenue,
                         avg_sale=avg_sale,
                         payment_data=payment_data,
                         hourly_data=hourly_data)

@bp.route('/sales/<sale_id>/details/html')
@login_required
def sale_details_html(sale_id):
    """Return sale details as HTML for modal"""
    # Query terpisah untuk sale details dengan semua data yang dibutuhkan
    sale = Sale.query.filter_by(
        id=sale_id,
        tenant_id=current_user.tenant_id
    ).options(
        selectinload(Sale.customer),
        selectinload(Sale.user)
    ).first_or_404()
    
    # Query items terpisah
    sale_items = SaleItem.query.filter_by(sale_id=sale_id)\
        .options(selectinload(SaleItem.product))\
        .all()
    
    # Convert timestamp to user timezone
    sale.local_created_at = convert_utc_to_user_timezone(sale.created_at)
    
    # Convert timestamps untuk items jika perlu
    for item in sale_items:
        if hasattr(item, 'created_at'):
            item.local_created_at = convert_utc_to_user_timezone(item.created_at)
    
    return render_template('reports/sale_details_modal.html', 
                         sale=sale, 
                         sale_items=sale_items)

@bp.route('/export-excel')
@login_required
def export_excel():
    """Export sales report to Excel - FIXED VERSION dengan local time"""
    # Get sales data dengan query sederhana
    sales = Sale.query.filter_by(tenant_id=current_user.tenant_id)\
        .options(selectinload(Sale.customer))\
        .order_by(Sale.created_at.desc()).all()
    
    # Convert timestamps ke local time untuk export
    for sale in sales:
        sale.local_created_at = convert_utc_to_user_timezone(sale.created_at)
    
    # Pre-calculate items count dengan query terpisah
    sale_ids = [sale.id for sale in sales]
    items_count_map = {}
    
    if sale_ids:
        items_count_query = db.session.query(
            SaleItem.sale_id,
            func.count(SaleItem.id).label('items_count')
        ).filter(SaleItem.sale_id.in_(sale_ids))\
         .group_by(SaleItem.sale_id).all()
        
        items_count_map = {sale_id: count for sale_id, count in items_count_query}
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Headers dengan styling
    headers = ['No Struk', 'Tanggal', 'Waktu', 'Customer', 'Jumlah Item', 'Total Amount', 'Metode Pembayaran']
    ws.append(headers)
    
    # Style header row
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=1, column=col).fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Data - gunakan local_created_at untuk tanggal dan waktu
    for sale in sales:
        customer_name = sale.customer.name if sale.customer else 'Walk-in'
        items_count = items_count_map.get(sale.id, 0)
        
        ws.append([
            sale.receipt_number,
            sale.local_created_at.strftime('%Y-%m-%d'),  # Gunakan local time
            sale.local_created_at.strftime('%H:%M'),     # Gunakan local time
            customer_name,
            items_count,
            sale.total_amount,
            sale.payment_method.upper() if sale.payment_method else 'UNKNOWN'
        ])
    
    # Auto-adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    
    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/export-pdf')
@login_required
def export_pdf():
    """Export sales report to PDF - FIXED VERSION dengan local time"""
    # Get sales data
    sales = Sale.query.filter_by(tenant_id=current_user.tenant_id)\
        .options(selectinload(Sale.customer))\
        .order_by(Sale.created_at.desc()).limit(50).all()
    
    # Convert timestamps ke local time untuk export
    for sale in sales:
        sale.local_created_at = convert_utc_to_user_timezone(sale.created_at)
    
    # Pre-calculate items count
    sale_ids = [sale.id for sale in sales]
    items_count_map = {}
    
    if sale_ids:
        items_count_query = db.session.query(
            SaleItem.sale_id,
            func.count(SaleItem.id).label('items_count')
        ).filter(SaleItem.sale_id.in_(sale_ids))\
         .group_by(SaleItem.sale_id).all()
        
        items_count_map = {sale_id: count for sale_id, count in items_count_query}
    
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    
    # Header
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 800, f"Laporan Penjualan - {current_user.tenant.name}")
    p.setFont("Helvetica", 10)
    p.drawString(100, 780, f"Dibuat pada: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    p.drawString(100, 765, f"Total Transaksi: {len(sales)}")
    
    # Table headers
    y_position = 740
    p.setFont("Helvetica-Bold", 10)
    p.drawString(50, y_position, "No Struk")
    p.drawString(120, y_position, "Tanggal")
    p.drawString(180, y_position, "Customer")
    p.drawString(280, y_position, "Items")
    p.drawString(320, y_position, "Total")
    p.drawString(380, y_position, "Pembayaran")
    
    # Data rows - gunakan local_created_at
    y_position -= 20
    p.setFont("Helvetica", 9)
    
    for sale in sales:
        if y_position < 100:  # New page jika perlu
            p.showPage()
            y_position = 740
            # Header untuk halaman baru
            p.setFont("Helvetica-Bold", 10)
            p.drawString(50, y_position, "No Struk")
            p.drawString(120, y_position, "Tanggal")
            p.drawString(180, y_position, "Customer")
            p.drawString(280, y_position, "Items")
            p.drawString(320, y_position, "Total")
            p.drawString(380, y_position, "Pembayaran")
            y_position -= 20
            p.setFont("Helvetica", 9)
        
        customer_name = sale.customer.name if sale.customer else 'Walk-in'
        items_count = items_count_map.get(sale.id, 0)
        
        p.drawString(50, y_position, sale.receipt_number)
        p.drawString(120, y_position, sale.local_created_at.strftime('%m/%d'))  # Local time
        p.drawString(180, y_position, customer_name[:15])  # Limit panjang nama
        p.drawString(280, y_position, str(items_count))
        p.drawString(320, y_position, f"Rp{sale.total_amount:,.0f}")
        p.drawString(380, y_position, sale.payment_method)
        y_position -= 15
    
    # Footer dengan total
    y_position -= 10
    p.setFont("Helvetica-Bold", 10)
    total_revenue = sum(sale.total_amount for sale in sales)
    p.drawString(300, y_position, f"TOTAL: Rp{total_revenue:,.0f}")
    
    p.save()
    buffer.seek(0)
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"sales_report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
        mimetype='application/pdf'
    )

@bp.route('/dashboard-data')
@login_required
def dashboard_data():
    """API data untuk dashboard charts dengan perhitungan yang benar"""
    from sqlalchemy import cast, Date
    
    # Gunakan UTC untuk konsistensi
    now_utc = datetime.utcnow()
    today_utc = now_utc.date()
    
    print(f"🔍 DEBUG: Today UTC = {today_utc}, Now UTC = {now_utc}")
    print(f"📅 DEBUG: Today weekday = {today_utc.weekday()} (0=Senin, 6=Minggu)")

    # 1. Data untuk chart (7 hari terakhir)
    start_date_chart = now_utc - timedelta(days=6)
    
    daily_sales = db.session.query(
        cast(Sale.created_at, Date).label('sale_date'),
        func.sum(Sale.total_amount),
        func.count(Sale.id)
    ).filter(
        Sale.tenant_id == current_user.tenant_id,
        Sale.created_at >= start_date_chart,
        Sale.created_at <= now_utc
    ).group_by(cast(Sale.created_at, Date)).order_by('sale_date').all()
    
    print(f"📊 DEBUG: Daily sales raw data = {daily_sales}")

    # 2. Data statistik yang benar
    
    # HARI INI - menggunakan UTC date
    today_sales = db.session.query(
        func.sum(Sale.total_amount),
        func.count(Sale.id)
    ).filter(
        Sale.tenant_id == current_user.tenant_id,
        cast(Sale.created_at, Date) == today_utc
    ).first()
    
    print(f"📅 DEBUG: Today sales = {today_sales}")

    # MINGGU INI - PERBAIKAN: Gunakan 7 hari terakhir termasuk hari ini
    start_of_week_utc = today_utc - timedelta(days=6)  # 7 hari termasuk hari ini
    end_of_week_utc = today_utc
    
    print(f"📅 DEBUG: Week range (7 days) = {start_of_week_utc} to {end_of_week_utc}")
    
    week_sales = db.session.query(
        func.sum(Sale.total_amount),
        func.count(Sale.id)
    ).filter(
        Sale.tenant_id == current_user.tenant_id,
        cast(Sale.created_at, Date) >= start_of_week_utc,
        cast(Sale.created_at, Date) <= end_of_week_utc
    ).first()
    
    print(f"📅 DEBUG: Week sales (7 days) = {week_sales}")

    # BULAN INI
    start_of_month_utc = today_utc.replace(day=1)
    next_month = start_of_month_utc.replace(day=28) + timedelta(days=4)
    end_of_month_utc = next_month - timedelta(days=next_month.day)
    
    month_sales = db.session.query(
        func.sum(Sale.total_amount),
        func.count(Sale.id)
    ).filter(
        Sale.tenant_id == current_user.tenant_id,
        cast(Sale.created_at, Date) >= start_of_month_utc,
        cast(Sale.created_at, Date) <= end_of_month_utc
    ).first()
    
    print(f"📅 DEBUG: Month sales = {month_sales}")

    # 3. Top products (30 hari terakhir)
    start_date_products = now_utc - timedelta(days=30)
    
    top_products = db.session.query(
        Product.name,
        func.sum(SaleItem.quantity),
        func.sum(SaleItem.total_price)
    ).join(SaleItem, Product.id == SaleItem.product_id)\
     .join(Sale, SaleItem.sale_id == Sale.id)\
     .filter(
         Sale.tenant_id == current_user.tenant_id,
         Sale.created_at >= start_date_products
     ).group_by(Product.id, Product.name)\
     .order_by(func.sum(SaleItem.total_price).desc())\
     .limit(10).all()

    # Format response dengan handling None values
    today_revenue = today_sales[0] if today_sales and today_sales[0] is not None else 0
    today_count = today_sales[1] if today_sales and today_sales[1] is not None else 0
    
    week_revenue = week_sales[0] if week_sales and week_sales[0] is not None else 0
    week_count = week_sales[1] if week_sales and week_sales[1] is not None else 0
    
    month_revenue = month_sales[0] if month_sales and month_sales[0] is not None else 0
    month_count = month_sales[1] if month_sales and month_sales[1] is not None else 0
    
    # Rata-rata transaksi - gunakan data minggu
    avg_sale_week = week_revenue / week_count if week_count > 0 else 0

    # Format daily sales dengan date sebagai string
    formatted_daily_sales = []
    
    # Generate complete 7-day range
    for i in range(7):
        current_date = (now_utc - timedelta(days=6-i)).date()
        
        # Cari data untuk tanggal ini
        daily_data = next((d for d in daily_sales if d[0] == current_date), None)
        
        formatted_daily_sales.append({
            'date': current_date.isoformat(),
            'revenue': float(daily_data[1] if daily_data else 0),
            'count': daily_data[2] if daily_data else 0
        })
    
    print(f"📈 DEBUG: Formatted daily sales = {formatted_daily_sales}")
    
    response_data = {
        'daily_sales': formatted_daily_sales,
        'top_products': [
            {
                'name': name, 
                'quantity': int(quantity or 0), 
                'revenue': float(revenue or 0)
            }
            for name, quantity, revenue in top_products
        ],
        'stats': {
            'today': {
                'revenue': float(today_revenue),
                'count': today_count
            },
            'week': {
                'revenue': float(week_revenue),
                'count': week_count
            },
            'month': {
                'revenue': float(month_revenue),
                'count': month_count
            },
            'avg_sale': float(avg_sale_week)
        }
    }
    
    print(f"📤 DEBUG: Final response data stats = {response_data['stats']}")
    
    return jsonify(response_data)

# Route untuk debug data sales dengan timezone info
@bp.route('/debug-sales-timezone')
@login_required
def debug_sales_timezone():
    """Debug route untuk melihat data sales dengan info timezone"""
    # Data sales terbaru
    recent_sales = Sale.query.filter_by(tenant_id=current_user.tenant_id)\
        .order_by(Sale.created_at.desc()).limit(5).all()
    
    debug_info = {
        'server_time_utc': datetime.utcnow().isoformat(),
        'server_time_local': datetime.now().isoformat(),
        'timezone_info': 'Testing timezone conversion',
        'recent_sales': []
    }
    
    for sale in recent_sales:
        local_time = convert_utc_to_user_timezone(sale.created_at)
        debug_info['recent_sales'].append({
            'id': sale.id,
            'receipt_number': sale.receipt_number,
            'created_at_utc': sale.created_at.isoformat(),
            'created_at_local': local_time.isoformat(),
            'total_amount': sale.total_amount,
            'payment_method': sale.payment_method
        })
    
    return jsonify(debug_info)